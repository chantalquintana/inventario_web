import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import pandas as pd
from PIL import Image, ImageTk, ImageDraw
import random
import string
import shutil
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
import gspread
import subprocess
import json
from openpyxl import load_workbook

def git_push_changes(mensaje_commit="Actualización inventario"):
    try:
        subprocess.run(["git", "add", "."], check=True)
        subprocess.run(["git", "commit", "-m", mensaje_commit], check=True)
        subprocess.run(["git", "push", "origin", "main"], check=True)
        print("Cambios subidos a GitHub correctamente.")
    except subprocess.CalledProcessError as e:
        print(f"Error al ejecutar git: {e}")

def exportar_a_json(df):
    productos = []
    for _, row in df.iterrows():
        productos.append({
            "codigo": row["Código"],
            "nombre": row["Nombre"],
            "descripcion": row["Descripción"],
            "precio_compra": row["Precio Compra"],
            "precio_venta": row["Precio Venta"],
            "stock": row["Stock"],
            "vendidos": row["Vendidos"],
            "ganancia": row["Ganancia"],
            "inversion": row["Inversión"],
            "imagen": f"imagenes/{row['Imagen']}" if row.get("Imagen") else ""
        })

    with open("productos.json", "w", encoding="utf-8") as f:
        json.dump(productos, f, ensure_ascii=False, indent=4)


CREDENCIALES_JSON = 'inventarioinfopar-d0cf52f91f49.json'
SPREADSHEET_ID = '1Cgo4C--ByZikIPyXvZJtnBsCjOM4W9fju_N3O9T-3V0'
SHEET_NAME = 'Inventario_Infopar'

RUTA_IMAGEN_PLACEHOLDER = "insumosfotocopiadoras.png"

SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']

class InventarioSheets:
    def __init__(self):
        creds = Credentials.from_service_account_file(CREDENCIALES_JSON, scopes=SCOPES)
        self.cliente = gspread.authorize(creds)
        self.hoja = self.cliente.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)

    def leer_datos(self):
        return self.hoja.get_all_records()

FILE_PATH = "inventario.xlsx"
IMG_FOLDER = "imagenes"

if not os.path.exists(IMG_FOLDER):
    os.makedirs(IMG_FOLDER)

def generar_codigo_unico(df):
    while True:
        codigo = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
        if codigo not in df["Código"].values:
            return codigo

def guardar_df(df):
    # Calcular Ganancia e Inversión antes de guardar
    df["Ganancia"] = (df["Precio Venta"] - df["Precio Compra"]) * df["Vendidos"]
    df["Inversión"] = df["Precio Compra"] * df["Stock"]

    # Convertir a enteros (sin coma decimal)
    df["Precio Compra"] = df["Precio Compra"].astype(int)
    df["Precio Venta"] = df["Precio Venta"].astype(int)
    df["Ganancia"] = df["Ganancia"].astype(int)
    df["Inversión"] = df["Inversión"].astype(int)

    df.to_excel(FILE_PATH, index=False)
    
    # Guardaste el DataFrame en 'inventario.xlsx'
    nombre_archivo = "inventario.xlsx"
    
    # Abrir archivo
    wb = load_workbook(nombre_archivo)
    ws = wb.active

    # Aplicar formato con separador de miles
    from openpyxl.styles import numbers
    for col in ["D", "E", "H", "I"]:  # D=Precio Compra, E=Precio Venta, H=Ganancia, I=Inversión
       for i in range(2, ws.max_row + 1):
         celda = ws[f"{col}{i}"]
         celda.number_format = "#,##0"  # Esto mostrará 1.000 en Excel

    # Encabezados de fórmulas (si no están ya)
    ws["H1"] = "Ganancia"
    ws["I1"] = "Inversión"

    # Desde la fila 2 hasta el final
    for i in range(2, ws.max_row + 1):
        # Ganancia = (Precio Venta - Precio Compra) * Vendidos
        ws[f"H{i}"] = f"=(E{i}-D{i})*G{i}"
    
        # Inversión = Precio Compra * Stock
        ws[f"I{i}"] = f"=D{i}*F{i}"

    # Guardar cambios
    wb.save(nombre_archivo)

    try:
        subir_df_a_sheets(df)
    except Exception as e:
        print(f"Error al subir datos a Google Sheets: {e}")

    try:
        exportar_a_json(df)
    except Exception as e:
        print(f"Error al exportar productos.json: {e}")

    git_push_changes("Actualización automática del inventario, Google Sheets y productos.json")

# --- FUNCIÓN PARA CARGAR LA IMAGEN DE PLACEHOLDER ---
def cargar_placeholder_personalizado(ruta_placeholder, size=(230, 230)):
    try:
        # Abre la imagen desde la ruta especificada
        img = Image.open(ruta_placeholder).convert("RGBA")

        try:
            resample = Image.Resampling.LANCZOS
        except AttributeError:
            resample = Image.LANCZOS

        # Escala la imagen para que encaje en el marco 230x230
        img.thumbnail(size, resample) 

        # Crear fondo blanco para centrar la imagen pequeña si es necesario
        fondo = Image.new("RGBA", size, (255, 255, 255, 255))
        x = (size[0] - img.width) // 2
        y = (size[1] - img.height) // 2
        fondo.paste(img, (x, y), img)

        return fondo.convert("RGB")
    except Exception as e:
        print(f"Error cargando imagen placeholder: {e}")
        # Retorna una imagen genérica si falla la carga (puedes omitir esto si quieres que falle si no encuentra la imagen)
        img_falla = Image.new("RGB", size, color='lightgray')
        return img_falla

def cargar_imagen_centrada(ruta, size=(230, 230), color_fondo=(255, 255, 255, 255)):
    try:
        img = Image.open(ruta).convert("RGBA")

        try:
            resample = Image.Resampling.LANCZOS
        except AttributeError:
            resample = Image.LANCZOS

        # Escalar manteniendo proporción
        img.thumbnail(size, resample)

        # Crear fondo blanco (o transparente) del tamaño deseado
        fondo = Image.new("RGBA", size, color_fondo)

        # Calcular posición para centrar
        x = (size[0] - img.width) // 2
        y = (size[1] - img.height) // 2

        # Pegar imagen en el fondo centrada
        fondo.paste(img, (x, y), img)

        return fondo.convert("RGB")  # Convertir a RGB para tkinter
    except Exception as e:
        print(f"Error cargando imagen: {e}")
        return None

def subir_df_a_sheets(df):
    creds = Credentials.from_service_account_file(CREDENCIALES_JSON, scopes=SCOPES)
    service = build('sheets', 'v4', credentials=creds)
    values = [df.columns.to_list()] + df.values.tolist()
    body = {'values': values}
    service.spreadsheets().values().update(
        spreadsheetId=SPREADSHEET_ID,
        range=f"{SHEET_NAME}!A1",
        valueInputOption="RAW",
        body=body
    ).execute()

class InventarioApp:
    def __init__(self, root):
        self.root = root
        self.root.title("INFOPAR PARAGUAY")
        ancho_ventana = 900
        alto_ventana = 580
        ancho_pantalla = self.root.winfo_screenwidth()
        alto_pantalla = self.root.winfo_screenheight()
        x = (ancho_pantalla - ancho_ventana) // 2
        y = (alto_pantalla - alto_ventana) // 2
        self.root.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")
        self.root.resizable(False, False)

        main_frame = tk.Frame(root)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        main_frame.grid_rowconfigure(0, weight=1)
        main_frame.grid_columnconfigure(0, weight=3)
        main_frame.grid_columnconfigure(1, weight=1)

        form_frame = tk.LabelFrame(main_frame, text="Datos del Producto", padx=10, pady=10)
        form_frame.grid(row=0, column=0, sticky="nsew")
        form_interno = tk.Frame(form_frame)
        form_interno.grid(row=0, column=0, columnspan=3)
        campos_frame = tk.Frame(form_interno)
        campos_frame.grid(row=0, column=0, columnspan=2, sticky="nw")

        labels = ["Nombre", "Descripción", "Precio Compra", "Precio Venta", "Stock", "Vendidos"]
        self.entries = {}
        for i, label in enumerate(labels):
            tk.Label(campos_frame, text=label, anchor="w", width=15).grid(row=i, column=0, sticky="w", pady=5, padx=(0,5))
            ent = tk.Entry(campos_frame, width=35)
            ent.grid(row=i, column=1, pady=5, sticky="w")
            self.entries[label] = ent

        tk.Label(campos_frame, text="Imagen", anchor="w", width=15).grid(row=len(labels), column=0, sticky="w", pady=5, padx=(0,5))
        self.imagen_path_var = tk.StringVar()
        self.lbl_imagen_path = tk.Label(campos_frame, textvariable=self.imagen_path_var, relief="sunken", width=35, anchor="w")
        self.lbl_imagen_path.grid(row=len(labels), column=1, sticky="w", pady=5)
        btn_sel_img = tk.Button(campos_frame, text="Seleccionar Imagen", command=self.seleccionar_imagen)
        btn_sel_img.grid(row=len(labels)+1, column=1, sticky="w", pady=(0,10))

        logo_frame = tk.Frame(form_interno, height=180, width=180)
        logo_frame.grid(row=0, column=2, rowspan=6, padx=(10,0), sticky="n")
        logo_frame.grid_propagate(False)

        logo_path = "logo_infopar.png"
        
        try:
            logo_img = Image.open(logo_path)
            logo_img.thumbnail((200, 200))
            self.logo_tk = ImageTk.PhotoImage(logo_img)
            lbl_logo = tk.Label(logo_frame, image=self.logo_tk)
            lbl_logo.pack(expand=True)
            logo_frame.grid_rowconfigure(0, weight=1)
            logo_frame.grid_columnconfigure(0, weight=1)
        except Exception:
            lbl_logo = tk.Label(logo_frame, text="Logo\nno disponible", relief="ridge", width=15, height=6)
            lbl_logo.pack(expand=True)

        btn_frame = tk.Frame(form_interno)
        btn_frame.grid(row=1, column=0, columnspan=3, pady=10, sticky="w")

        self.btn_agregar = tk.Button(btn_frame, text="Agregar", width=12, command=self.agregar_producto)
        self.btn_agregar.grid(row=0, column=0, padx=(0,5))
        self.btn_editar = tk.Button(btn_frame, text="Editar", width=12, command=self.editar_producto, state="disabled")
        self.btn_editar.grid(row=0, column=1, padx=5)
        self.btn_eliminar = tk.Button(btn_frame, text="Eliminar", width=12, command=self.eliminar_producto, state="disabled")
        self.btn_eliminar.grid(row=0, column=2, padx=5)
        self.btn_limpiar = tk.Button(btn_frame, text="Limpiar", width=12, command=self.limpiar_campos)
        self.btn_limpiar.grid(row=0, column=3, padx=5)

        self.right_frame = tk.Frame(main_frame, width=240, height=240)
        self.right_frame.grid(row=0, column=1, sticky="nsew", padx=10)
        self.right_frame.grid_propagate(False)

        self.img_label = tk.Label(self.right_frame, text="Imagen del Producto", relief="groove")
        self.img_label.place(relx=0.5, rely=0.5, anchor="center", width=230, height=230)

        # *** Carga tu imagen específica desde tu PC como placeholder ***
        img_placeholder_pil = cargar_placeholder_personalizado(RUTA_IMAGEN_PLACEHOLDER)
        self.img_generica_tk = ImageTk.PhotoImage(img_placeholder_pil)
        
        # Muestra el placeholder inmediatamente (opcional)
        self.img_label.config(image=self.img_generica_tk, text="")

        bottom_frame = tk.Frame(root)
        bottom_frame.pack(fill="both", expand=True, padx=10, pady=10)

        search_frame = tk.LabelFrame(bottom_frame, text="Buscar Producto", padx=10, pady=10)
        search_frame.pack(fill="x", pady=(0,10))

        tk.Label(search_frame, text="Nombre o descripción:").pack(side="left")
        self.entry_buscar = tk.Entry(search_frame, width=30)
        self.entry_buscar.pack(side="left", padx=5)
        btn_buscar = tk.Button(search_frame, text="Buscar", command=self.buscar_producto)
        btn_buscar.pack(side="left", padx=5)
        btn_mostrar = tk.Button(search_frame, text="Mostrar Todo", command=lambda: self.llenar_tabla(pd.DataFrame(self.inventario_sheets.leer_datos())))
        btn_mostrar.pack(side="left", padx=5)

        table_frame = tk.Frame(bottom_frame)
        table_frame.pack(fill="both", expand=True)

        columnas = ("Código", "Nombre", "Descripción", "Precio Compra", "Precio Venta", "Stock", "Vendidos", "Ganancia", "Inversión")
        self.tree = ttk.Treeview(table_frame, columns=columnas, show="headings", height=10)

        for col in columnas:
            if col == "Código":
                ancho = 70
                anchor = "w"
            elif col == "Nombre":
                ancho = 150
                anchor = "w"
            elif col == "Descripción":
                ancho = 250
                anchor = "w"
            elif col in ("Ganancia", "Inversión"):
                ancho = 90
                anchor = "e"
            else:
                ancho = 70
                anchor = "e"

            self.tree.heading(col, text=col)
            self.tree.column(col, width=ancho, anchor=anchor)

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        vsb.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=vsb.set)

        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        hsb.pack(side="bottom", fill="x")
        self.tree.configure(xscrollcommand=hsb.set)

        self.tree.pack(fill="both", expand=True)

        self.tree.bind("<Button-1>", self.bloquear_redimension_columnas)

        vista_frame = tk.LabelFrame(bottom_frame, text="Vista Ampliada", padx=10, pady=10)
        vista_frame.pack(fill="both", expand=False, pady=(5, 0))

        self.lbl_nombre_ampliado = tk.Label(vista_frame, text="Nombre: ", font=("Arial", 12, "bold"), anchor="w", justify="left")
        self.lbl_nombre_ampliado.pack(fill="x", pady=2)

        txt_scroll_frame = tk.Frame(vista_frame)
        txt_scroll_frame.pack(fill="both", expand=True)

        self.txt_descripcion_ampliada = tk.Text(txt_scroll_frame, height=5, font=("Arial", 11), wrap="word")
        self.txt_descripcion_ampliada.pack(side="left", fill="both", expand=True)

        scroll = tk.Scrollbar(txt_scroll_frame, command=self.txt_descripcion_ampliada.yview)
        scroll.pack(side="right", fill="y")

        self.txt_descripcion_ampliada.configure(yscrollcommand=scroll.set)
        self.txt_descripcion_ampliada.config(state="disabled")

        self.tree.bind("<<TreeviewSelect>>", self.mostrar_detalle_producto)
        self.tree.bind("<<TreeviewSelect>>", self.on_fila_seleccionada)

        self.producto_seleccionado_codigo = None

        self.inventario_sheets = InventarioSheets()
        self.df = pd.DataFrame()  # Se asignará en la función de sincronización

        self.preguntar_actualizar_desde_sheets()

        self.limpiar_campos()

        self.actualizar_periodicamente()

        self.imagen_actual = None

    def preguntar_actualizar_desde_sheets(self):
        try:
            # Leer datos locales si existen
            if os.path.exists(FILE_PATH):
                df_local = pd.read_excel(FILE_PATH)
            else:
                df_local = pd.DataFrame()

            # Leer datos desde Sheets
            df_sheets = pd.DataFrame(self.inventario_sheets.leer_datos())

            # Normalizar tipos de datos para comparación (convertir todo a str)
            df_local_str = df_local.astype(str).fillna("")
            df_sheets_str = df_sheets.astype(str).fillna("")

            # Comparar si son exactamente iguales
            iguales = df_local_str.equals(df_sheets_str)

            if not iguales:
                # Actualizar local con datos de Sheets
                self.df = df_sheets
                guardar_df(self.df)
                self.llenar_tabla(self.df)
                messagebox.showinfo("Actualizado", "Inventario local actualizado desde Google Sheets.")
            else:
                if not df_local.empty:
                    self.df = df_local
                    self.llenar_tabla(self.df)
        except Exception as e:
            print(f"Error sincronizando inventario con Sheets: {e}")
            # Intentar cargar local si existe
            if os.path.exists(FILE_PATH):
                self.df = pd.read_excel(FILE_PATH)
                self.llenar_tabla(self.df)

    def actualizar_periodicamente(self):
        self.preguntar_actualizar_desde_sheets()
        # Actualiza cada 5 minutos (300000 ms)
        self.root.after(300000, self.actualizar_periodicamente)

    def bloquear_redimension_columnas(self, event):
        # Prevenir que el usuario cambie ancho columnas
        if self.tree.identify_region(event.x, event.y) == "separator":
            return "break"

        
    def mostrar_detalle_producto(self, event):
        seleccionado = self.tree.selection()
        if seleccionado:
            item = self.tree.item(seleccionado[0])
            vals = item["values"]
            self.lbl_nombre_ampliado.config(text=f"Nombre: {vals[1]}")
            self.txt_descripcion_ampliada.config(state="normal")
            self.txt_descripcion_ampliada.delete(1.0, "end")
            self.txt_descripcion_ampliada.insert("end", vals[2])
            self.txt_descripcion_ampliada.config(state="disabled")

            # Mostrar imagen del producto
            img_nombre = self.df.loc[self.df["Código"] == vals[0], "Imagen"].values
            if len(img_nombre) > 0 and img_nombre[0]:
                ruta_img = os.path.join(IMG_FOLDER, img_nombre[0])
                if os.path.exists(ruta_img):
                    img_pil = cargar_imagen_centrada(ruta_img)
                    if img_pil:
                        self.imagen_actual = ImageTk.PhotoImage(img_pil)
                        self.img_label.config(image=self.imagen_actual, text="")
                    else:
                        self.img_label.config(image=self.img_generica_tk, text="")
                else:
                    self.img_label.config(image=self.img_generica_tk, text="")
            else:
                self.img_label.config(image=self.img_generica_tk, text="")


            # Activar botones editar y eliminar
            self.btn_editar.config(state="normal")
            self.btn_eliminar.config(state="normal")

    def llenar_tabla(self, df):
        self.tree.delete(*self.tree.get_children())

        # Calcular Ganancia e Inversión para mostrar en tabla
        df = df.copy()
        if "Ganancia" not in df.columns or "Inversión" not in df.columns:
            df["Ganancia"] = (df["Precio Venta"] - df["Precio Compra"]) * df["Vendidos"]
            df["Inversión"] = df["Precio Compra"] * df["Stock"]

        for _, row in df.iterrows():
            valores = (
                str(row["Código"]),
                row["Nombre"],
                row["Descripción"],
                f"{int(row['Precio Compra']):,}".replace(",", "."),
                f"{int(row['Precio Venta']):,}".replace(",", "."),
                f"{int(row['Stock'])}",
                f"{int(row['Vendidos'])}",
                f"{int(row['Ganancia']):,}".replace(",", "."),
                f"{int(row['Inversión']):,}".replace(",", ".")
            )
            self.tree.insert("", "end", values=valores)
            
        self.df = df  # Actualizar df interno

    def limpiar_campos(self):
        for ent in self.entries.values():
            ent.delete(0, "end")
        self.imagen_path_var.set("")
        self.btn_agregar.config(state="normal")
        self.btn_editar.config(state="disabled")
        self.btn_eliminar.config(state="disabled")
        self.tree.selection_remove(self.tree.selection())
        self.img_label.config(image=self.img_generica_tk, text="Imagen del Producto")

    def seleccionar_imagen(self):
        ruta = filedialog.askopenfilename(title="Seleccionar imagen", filetypes=[("Archivos de imagen", "*.jpg;*.jpeg;*.png;*.bmp")])
        if ruta:
            self.imagen_path_var.set(ruta)

    def validar_campos(self):
        try:
            nombre = self.entries["Nombre"].get().strip()
            descripcion = self.entries["Descripción"].get().strip()

            precio_compra_str = self.entries["Precio Compra"].get().strip()
            precio_venta_str = self.entries["Precio Venta"].get().strip()
            stock_str = self.entries["Stock"].get().strip()
            vendidos_str = self.entries["Vendidos"].get().strip()

            if not precio_compra_str or not precio_venta_str or not stock_str or not vendidos_str:
              messagebox.showerror("Error", "Los campos numéricos no pueden estar vacíos.")
              return None
            
            precio_compra = int(self.entries["Precio Compra"].get())
            precio_venta = int(self.entries["Precio Venta"].get())
            stock = int(self.entries["Stock"].get())
            vendidos = int(self.entries["Vendidos"].get())

            if not nombre:
                messagebox.showerror("Error", "El campo Nombre es obligatorio.")
                return None
            if precio_compra < 0 or precio_venta < 0 or stock < 0 or vendidos < 0:
                messagebox.showerror("Error", "Los valores numéricos no pueden ser negativos.")
                return None
            if precio_venta < precio_compra:
                messagebox.showwarning("Advertencia", "El Precio de Venta es menor que el Precio de Compra.")
            return nombre, descripcion, precio_compra, precio_venta, stock, vendidos
        except ValueError:
            messagebox.showerror("Error", "Por favor, ingrese valores numéricos válidos en Precio, Stock y Vendidos.")
            return None

    def copiar_imagen(self, ruta_imagen):
        if not ruta_imagen:
            return ""
        try:
            nombre_archivo = os.path.basename(ruta_imagen)
            destino = os.path.join(IMG_FOLDER, nombre_archivo)

            if not os.path.exists(IMG_FOLDER):
              os.makedirs(IMG_FOLDER)

            if not os.path.exists(destino):
              shutil.copy(ruta_imagen, destino)

            return nombre_archivo
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo copiar la imagen: {e}")
        return ""

    def on_fila_seleccionada(self, event):
        selected = self.tree.selection()
        if selected:
            item = self.tree.item(selected[0])
            codigo = item["values"][0]
            producto = self.df[self.df["Código"] == codigo].iloc[0]
            self.producto_seleccionado_codigo = codigo

            for label in self.entries:
                valor = producto[label]
                if pd.isna(valor):
                    valor = ""
                self.entries[label].delete(0, tk.END)
                self.entries[label].insert(0, str(valor))

            nombre_img = producto.get("Imagen", None)
            if nombre_img:
                ruta_img = os.path.join(IMG_FOLDER, nombre_img)
                if os.path.exists(ruta_img):
                    img_pil = cargar_imagen_centrada(ruta_img)
                    if img_pil:
                        self.imagen_producto_actual = ImageTk.PhotoImage(img_pil)
                        self.img_label.config(image=self.imagen_producto_actual, text="")
                        self.img_label.image = self.imagen_producto_actual
                    else:
                        self.img_label.config(image=self.img_generica_tk, text="")
                        self.img_label.image = self.img_generica_tk
                else:
                    self.img_label.config(image=self.img_generica_tk, text="")
                    self.img_label.image = self.img_generica_tk
            else:
                self.img_label.config(image=self.img_generica_tk, text="")
                self.img_label.image = self.img_generica_tk


            self.imagen_path_var.set("")
            self.btn_editar.config(state="normal")
            self.btn_eliminar.config(state="normal")
            self.btn_agregar.config(state="disabled")

            self.lbl_nombre_ampliado.config(text=f"Nombre: {producto['Nombre']}")
            self.txt_descripcion_ampliada.config(state="normal")
            self.txt_descripcion_ampliada.delete(1.0, tk.END)
            self.txt_descripcion_ampliada.insert(tk.END, producto["Descripción"])
            self.txt_descripcion_ampliada.config(state="disabled")


    def agregar_producto(self):
        validacion = self.validar_campos()
        if validacion is None:
            return
        nombre, descripcion, precio_compra, precio_venta, stock, vendidos = validacion
        imagen_nombre = self.copiar_imagen(self.imagen_path_var.get())

        if self.df.empty:
            self.df = pd.DataFrame(columns=["Código", "Nombre", "Descripción", "Precio Compra", "Precio Venta", "Stock", "Vendidos", "Imagen", "Ganancia", "Inversión"])

        codigo = generar_codigo_unico(self.df)

        nuevo_producto = {
            "Código": codigo,
            "Nombre": nombre,
            "Descripción": descripcion,
            "Precio Compra": precio_compra,
            "Precio Venta": precio_venta,
            "Stock": stock,
            "Vendidos": vendidos,
            "Imagen": imagen_nombre
        }

        nuevo_df = pd.DataFrame([nuevo_producto])  # Convierte dict a DF con 1 fila
        self.df = pd.concat([self.df, nuevo_df], ignore_index=True)

        guardar_df(self.df)
        self.llenar_tabla(self.df)
        self.limpiar_campos()
        messagebox.showinfo("Éxito", "Producto agregado correctamente.")

    def editar_producto(self):
        seleccionado = self.tree.selection()
        if not seleccionado:
            messagebox.showwarning("Advertencia", "Seleccione un producto para editar.")
            return
        
        validacion = self.validar_campos()
        if validacion is None:
            return
        nombre, descripcion, precio_compra, precio_venta, stock, vendidos = validacion
        imagen_nombre = self.copiar_imagen(self.imagen_path_var.get())

        codigo_seleccionado = self.tree.item(seleccionado[0])["values"][0]

        #Obtener el índice del producto en el DataFrame
        idx = self.df.index[self.df["Código"] == codigo_seleccionado].tolist()
        if not idx:
            messagebox.showerror("Error", "Producto no encontrado en el inventario.")
            return
        idx = idx[0]# Tomar el primer (y único) índice

        # Actualizar datos del producto en el DataFrame
        self.df.at[idx, "Nombre"] = nombre
        self.df.at[idx, "Descripción"] = descripcion
        self.df.at[idx, "Precio Compra"] = precio_compra
        self.df.at[idx, "Precio Venta"] = precio_venta
        self.df.at[idx, "Stock"] = stock
        self.df.at[idx, "Vendidos"] = vendidos
        
        if imagen_nombre:
            self.df.at[idx, "Imagen"] = imagen_nombre

        # Manejo de la imagen: Si se seleccionó una nueva imagen, se copia y se actualiza el nombre en el DF
        if self.imagen_path_var.get():
            ruta_imagen_original = self.imagen_path_var.get()
            ext = os.path.splitext(ruta_imagen_original)[1]
            nombre_imagen_nuevo = f"{codigo_seleccionado}{ext}" # Usar el código como nombre de archivo para la imagen
            destino = os.path.join(IMG_FOLDER, nombre_imagen_nuevo)

            try:
                # Si ya existe una imagen con ese nombre, la eliminamos primero para no tener duplicados
                # Esto es importante si el tipo de archivo cambió (ej: de .png a .jpg)
                imagen_anterior = self.df.at[idx, "Imagen"]
                if imagen_anterior and os.path.exists(os.path.join(IMG_FOLDER, imagen_anterior)):
                    if imagen_anterior != nombre_imagen_nuevo: # Evitar borrar si es la misma imagen o mismo nombre
                        try:
                            os.remove(os.path.join(IMG_FOLDER, imagen_anterior))
                        except Exception as e:
                            print(f"Advertencia: No se pudo eliminar la imagen anterior '{imagen_anterior}': {e}")
                shutil.copy2(ruta_imagen_original, destino)
                self.df.at[idx, "Imagen"] = nombre_imagen_nuevo
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo copiar la nueva imagen: {e}")
                return

        guardar_df(self.df)# Guarda el DataFrame actualizado en Excel y Sheets
        self.llenar_tabla(self.df)
        self.limpiar_campos()
        messagebox.showinfo("Éxito", "Producto editado correctamente.")

    def eliminar_producto(self):
        seleccionado = self.tree.selection()
        if not seleccionado:
            messagebox.showwarning("Advertencia", "Seleccione un producto para eliminar.")
            return
        codigo = self.tree.item(seleccionado[0])["values"][0]

        idx = self.df.index[self.df["Código"] == codigo].tolist()
        if not idx:
            messagebox.showerror("Error", "Producto no encontrado en el inventario.")
            return
        idx = idx[0]

        resp = messagebox.askyesno("Confirmar", f"¿Eliminar el producto '{self.df.at[idx, 'Nombre']}'?")
        if resp:
            # Opcional: borrar imagen del disco (solo si no está usada en otro producto)
            nombre_img = self.df.at[idx, "Imagen"]
            if nombre_img:
                otros = self.df[self.df["Imagen"] == nombre_img]
                if len(otros) <= 1:
                    try:
                        os.remove(os.path.join(IMG_FOLDER, nombre_img))
                    except Exception:
                        pass

            self.df = self.df.drop(idx).reset_index(drop=True)
            guardar_df(self.df)
            self.llenar_tabla(self.df)
            self.limpiar_campos()
            messagebox.showinfo("Éxito", "Producto eliminado correctamente.")

    def buscar_producto(self):
        texto = self.entry_buscar.get().strip().lower()
        if texto == "":
            self.llenar_tabla(self.df)
            return

        df_filtrado = self.df[self.df["Nombre"].str.lower().str.contains(texto) | self.df["Descripción"].str.lower().str.contains(texto)]
        if df_filtrado.empty:
            messagebox.showinfo("Buscar", "No se encontraron productos que coincidan.")
            return
        self.llenar_tabla(df_filtrado)

if __name__ == "__main__":
    import threading
    from flask import Flask, request, jsonify, send_from_directory
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials

    # --- Configuración Google Sheets ---
    scope = ["https://spreadsheets.google.com/feeds","https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name("credenciales.json", scope)
    client = gspread.authorize(creds)
    sheet = client.open("inventario_infopar").sheet1

    # --- Crear servidor Flask ---
    app = Flask(__name__, static_folder='.')  # '.' porque tus archivos están en la misma carpeta

    # Servir la página principal
    @app.route('/')
    def index():
        return send_from_directory('.', 'index.html')  # abre tu HTML principal

    # Servir cualquier archivo estático (imagenes, JS, CSS)
    @app.route('/<path:filename>')
    def static_files(filename):
        return send_from_directory('.', filename)

    # Ruta para actualizar stock
    @app.route("/actualizar_stock", methods=["POST"])
    def actualizar_stock():
        data = request.get_json()
        codigo = data.get("codigo")
        nuevo_stock = data.get("nuevoStock")
        all_codes = sheet.col_values(1)
        if codigo in all_codes:
            fila = all_codes.index(codigo) + 1
            sheet.update_cell(fila, 6, str(nuevo_stock))
            return jsonify({"status":"ok","fila":fila,"nuevoStock":nuevo_stock})
        else:
            return jsonify({"status":"error","message":"Código no encontrado"}), 404

    # --- Iniciar Flask ---
    app.run(host="0.0.0.0", port=5000)
